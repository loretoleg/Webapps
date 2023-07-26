import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import datetime
import os
import sys
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import json
import requests

def lambda_handler(event, context):                                                  
                                               
  # Get the body from the event
  body = event['body']                                                        

  start_index = body.find("{")
  end_index = body.rfind("}")
  json_data = body[start_index:end_index+1]
  dict_data = json.loads(json_data)
  
  # show the browse dialog box and get the selected file path
  file_path = '/tmp/temp_excel.xlsx'    
 
  rec_email = dict_data['q9_email']
  num_cist = int(dict_data['q5_numeroDe'])
  opt_var = dict_data['q7_tipoDe']
  url_datos = dict_data['archivoExcel'][0]

  headers = {'Authorization': os.environ["jot_token"]}
  response = requests.get(url_datos, headers=headers)

  file_path = '/tmp/temp_excel.xlsx'

  if response.status_code == 200:
    # Create a temporary file to save the downloaded content
    with open(file_path, 'wb') as f:
      f.write(response.content)
  else: print("Error downloading the file.")
################################################################################
  Comp_A = os.environ["Comp_A"]
  Comp_B = os.environ["Comp_B"]
################################################################################

  def max_num_pages(df):
    num_rows = len(df)
    if num_rows <= 34: return 1
    else:
        num_pages = (num_rows - 34) // 35
        if (num_rows - 34) % 35 == 0: return num_pages + 1
        else: return num_pages + 2

  def to_spa(x):
      if x < 0:
          sign = "-"
          x = -x
      else:
          sign = ""
      return sign + "{:,.2f}".format(x).replace(',', '-').replace('.', ',').replace('-', '.')

  def create_report(df, num_cisterna):

      # Cantidad Recibida por 'Company A'
      rec_psa = df[(df['CISTERNA']==num_cisterna) & (df['TRANSACCIÓN']=='Recibido')]\
      .groupby('FECHA').sum()['CANTIDAD LITROS'].to_frame(name=f'Cantidad Recibida por {Comp_A}')

      # Cantidad Despachada por Personal de 'Company A' (A)
      dsp_psa = df[(df['CISTERNA']==num_cisterna) & (df['TRANSACCIÓN']=='Despachado')\
                  & (df['BENEFICIARIO']!=f'{Comp_B}')].groupby('FECHA')\
                  .sum()['CANTIDAD LITROS'].to_frame(name=f'Cantidad Despachada por Personal de {Comp_A} (A)')


      # Cantidad Despachada por Personal de 'Company B' (B)
      dsp_psb = df[(df['CISTERNA']==num_cisterna) & (df['TRANSACCIÓN']=='Despachado')\
                  & (df['BENEFICIARIO']==f'{Comp_B}')].groupby('FECHA').sum()['CANTIDAD LITROS']\
                  .to_frame(name=f'Cantidad Despachada por Personal de {Comp_B} (B)')

      # Create Report DF
      df1 = pd.concat([rec_psa, dsp_psa, dsp_psb], axis=1).reset_index()
      df1 = df1.set_index('FECHA')
      df1 = df1.resample('D').asfreq()
      df1.fillna(0, inplace=True)

      # Create a new range index & Assign the new range index to the index of the DataFrame
      df1 = df1.reset_index().set_index(pd.RangeIndex(start=1, stop=len(df1)+1, name="Items"))

      #Format Fecha
      df1['FECHA'] = df1['FECHA'].dt.strftime('%d/%m/%Y')

      #Add Lugar de Despacho o Recibo
      df1['Lugar de Despacho o Recibo'] = os.environ["Site_A"]

      #Add N° de Cisterna Column
      df1['N° de Cisterna'] = f'Cisterna {num_cisterna}'

      #Add Total Cantidad Despachada
      df1['Total Cantidad Despachada (A+B)'] = df1[f'Cantidad Despachada por Personal de {Comp_A} (A)'] +\
      df1[f'Cantidad Despachada por Personal de {Comp_B} (B)']

      #Add Balance General del Cisterna
      # set the initial value of the 'Balance General del Cisterna' column
      df1.loc[1, 'Balance General del Cisterna'] = df1.loc[1,f'Cantidad Recibida por {Comp_A}'] \
          - df1.loc[1,'Total Cantidad Despachada (A+B)']

      df1.loc[2:, 'Balance General del Cisterna'] = round(df1.loc[1, 'Balance General del Cisterna']\
      - np.cumsum(df1.loc[2:, 'Total Cantidad Despachada (A+B)']), 2)

      #Add Observación. Column
      for index, row in df1.iterrows():
        num_factura = df[(df['CISTERNA'] == num_cisterna) & (df['TRANSACCIÓN'] == 'Recibido')]["N° FACTURA"].iloc[0]
        recibido = row[f'Cantidad Recibida por {Comp_A}']
        despachado = row['Total Cantidad Despachada (A+B)']
        restante = row['Balance General del Cisterna']
        if despachado == 0:
            observation = 'No Hay Litros Consumidos'
        else:
            observation = f'Cisterna {num_cisterna}: N° de Control: {num_factura}. Recibido: {to_spa(recibido)}. '+\
            f'Despachado: {to_spa(despachado)}. Quedan: {to_spa(restante)}.'
        df1.at[index, 'Observación.'] = observation

      #Rearrange columns
      df1 = df1.reindex(columns=['FECHA', 'N° de Cisterna', 'Lugar de Despacho o Recibo', f'Cantidad Recibida por {Comp_A}',
                                f'Cantidad Despachada por Personal de {Comp_A} (A)',
                                f'Cantidad Despachada por Personal de {Comp_B} (B)',
                                'Total Cantidad Despachada (A+B)', 'Balance General del Cisterna', 'Observación.' ])
      
      #More math before conversion
      total_rec = to_spa(df1[f'Cantidad Recibida por {Comp_A}'].sum())
      total_desp = to_spa(df1['Total Cantidad Despachada (A+B)'].sum())
      disp_mes = to_spa(df1['Balance General del Cisterna'].iloc[-1])
      who = 'PSPSA.' if "-" in disp_mes else 'la División.'
      observacion = (f'Total Recibido: {total_rec} Litros; '
                    f'Total Despachado: {total_desp} Litros; '
                    f'\nQuedaron del Mes: {disp_mes} Litros. '
                    f'Consumidos Por {who}')

      #Convert to string to fit format
      df1[f'Cantidad Recibida por {Comp_A}'] = df1[f'Cantidad Recibida por {Comp_A}'].apply(to_spa)
      df1[f'Cantidad Despachada por Personal de {Comp_A} (A)'] = df1[f'Cantidad Despachada por Personal de {Comp_A} (A)'].apply(to_spa)
      df1[f'Cantidad Despachada por Personal de {Comp_B} (B)'] = df1[f'Cantidad Despachada por Personal de {Comp_B} (B)'].apply(to_spa)
      df1['Total Cantidad Despachada (A+B)'] = df1['Total Cantidad Despachada (A+B)'].apply(to_spa)
      df1['Balance General del Cisterna'] = df1['Balance General del Cisterna'].apply(to_spa)

      return df1, total_rec, total_desp, disp_mes, observacion

  #Part 3 (openpyxl):
  def page_num(pag_num, max_page):
    return f'Pagina {pag_num} de {max_page}'

  def copy_format(range1, range2):
    # iterate over the rows of the first range
    for row1, row2 in zip(ws[range1], ws[range2]):
      # iterate over the cells of the row
      for cell1, cell2 in zip(row1, row2):
          # copy the font of the cell from the first range to the corresponding cell in the second range
          cell2.font = Font(name=cell1.font.name, size=cell1.font.size,
                            bold=cell1.font.bold, italic=cell1.font.italic,
                            underline=cell1.font.underline)
          cell2.fill = PatternFill(start_color=cell1.fill.start_color,
                                  end_color=cell1.fill.end_color,
                                  fill_type=cell1.fill.fill_type)
          cell2.alignment = Alignment(horizontal=cell1.alignment.horizontal,
                                      vertical=cell1.alignment.vertical,
                                      wrap_text=cell1.alignment.wrap_text)
          row_num2 = cell2.row
          row_num1 = cell1.row
          ws.row_dimensions[row_num2].height = ws.row_dimensions[row_num1].height
          border = Border(top=Side(style=cell1.border.top.style,color=cell1.border.top.color),
                        left=Side(style=cell1.border.left.style,color=cell1.border.left.color),
                        right=Side(style=cell1.border.right.style,color=cell1.border.right.color),
                        bottom=Side(style=cell1.border.bottom.style,color=cell1.border.bottom.color))
          cell2.border = border

  def choice(variable):
      choice = variable
      #Link visual (key) to Sheet name (value)
      option_map = {'REPORTE DIARIO GASOLINA': os.environ["Sheet_1"],                 
                    'REPORTE DIARIO DIESEL': os.environ["Sheet_2"]}                   
      selected_option = option_map[choice]
      return selected_option




  var=opt_var                                                                     

  # get the selected sheet name from the dropdown
  sheet_name = choice(var)                                                        

  # read the Excel file into a pandas DataFrame
  df = pd.read_excel(file_path, sheet_name=sheet_name)
  os.remove(file_path)                                                                                  
  df = df.rename(columns={df.columns[1]: 'FECHA'})
  num_cisterna = int(num_cist)                                                     

  df, total_rec, total_desp, disp_mes, observacion = create_report(df,num_cisterna)


  ###################################################################################################
  ###################################################################################################
  ###################################################################################################


  #today date
  today = datetime.datetime.today().strftime('%d/%m/%Y')

  #Creation Variables
  max_page = max_num_pages(df)
  df.insert(3, 'Empty', '') #Insert empty column for merging

  # create an openpyxl workbook and select the active worksheet
  wb = load_workbook('Template.xlsx')                                              
  ws = wb.active
  max_temp = ws.max_row #max template row

  #FORMAT TEMPLATE BEFORE COPY

  ws['A1'] = f'{os.environ["Head_1"]}\nBalance de Combustible Resumen Diario por Cisterna' 
  ws['F4'] = f'Cantidad Recibida por {Comp_A}'
  ws['G4'] = f'Cantidad Despachada por Personal de {Comp_A}\n(A)'
  ws['H4'] = f'Cantidad Despachada Sin Personal de {Comp_A}\n(B)'
  ws['C41'] = os.environ["nom_c"]
  ws['E41'] = os.environ["nom_l"]
  ws['I41'] = os.environ["nom_d"]
  ws['K41'] = os.environ["nom_i"]
  ws['L41'] = os.environ["nom_j"]
  ws['C42'] = os.environ["ced_c"]
  ws['E42'] = os.environ["ced_l"]
  ws['I42'] = os.environ["ced_d"]
  ws['K42'] = os.environ["ced_i"]
  ws['L42'] = os.environ["ced_j"]
  ws['C43'] = os.environ["cargo_c"]
  ws['E43'] = os.environ["cargo_l"]
  ws['I43'] = os.environ["cargo_b_d"]
  ws['K43'] = os.environ["cargo_i"]
  ws['L43'] = os.environ["cargo_j"]
  ws['A44'] = os.environ["Foot_1"]
  ws['G44'] = os.environ["Foot_2"]


  ###############################################################################
  #Modify date in template
  ws['K3'] = today

  #Modify Page Number
  ws['L3'] = page_num(1, max_page)

  #Modify Combustible
  ws['K1'] = sheet_name.split("_")[0]

  #Modify Footer Supervisor
  if sheet_name.split("_")[0] == 'GASOIL':
    ws[f'I{max_temp-3}'] = f'{os.environ["nom_b"]}'
    ws[f'I{max_temp-2}'] = f'{os.environ["ced_b"]}'

  #Modify Footer Last Page
  ws[f'C{max_temp-5}'] = total_rec
  ws[f'F{max_temp-5}'] = total_desp
  ws[f'I{max_temp-5}'] = disp_mes
  ws[f'J{max_temp-5}'] = observacion

  #Add footer total row alignment
  ali = Alignment(vertical = 'center', horizontal = 'right')
  ws[f'C{max_temp-5}'].alignment = ali
  ws[f'F{max_temp-5}'].alignment = ali
  ws[f'I{max_temp-5}'].alignment = ali

  # copy template for more pages
  for page in range(1,max_page):
    shift = max_temp*page

    # Copy rows 1 to 44 from the Template
    for row in ws.iter_rows(min_row=1, max_row=max_temp, values_only=True):
      ws.append(row)

    #Add Page Number
    ws[f'L{3+shift}'] = page_num(page+1, max_page)

    #COPY FONT FROM PAGE 1 TO PAGE X
    copy_format('A1:L4',f'A{(1+shift)}:L{(4+shift)}') #header
    copy_format(f'A{max_temp-5}:L{max_temp}', f'A{(max_temp-5)+shift}:L{max_temp+shift}') #footer

    #MERGING
    #HEADER FORMAT
    ws.merge_cells(range_string=f'A{1+shift}:I{3+shift}')
    ws.merge_cells(range_string=f'J{1+shift}:J{2+shift}')
    ws.merge_cells(range_string=f'K{1+shift}:L{2+shift}')
    ws.merge_cells(range_string=f'D{4+shift}:E{4+shift}')
    ws.merge_cells(range_string=f'K{4+shift}:L{4+shift}')

    #FOOTER FORMAT
    ws.merge_cells(range_string=f'A{(max_temp-5)+shift}:B{(max_temp-5)+shift}')
    ws.merge_cells(range_string=f'D{(max_temp-5)+shift}:E{(max_temp-5)+shift}')
    ws.merge_cells(range_string=f'G{(max_temp-5)+shift}:H{(max_temp-5)+shift}')
    ws.merge_cells(range_string=f'J{(max_temp-5)+shift}:L{(max_temp-5)+shift}')

    #40-43
    # Define the row and column ranges
    row_range = range((max_temp-4)+shift, max_temp+shift)
    letter_range = ['A', 'C', 'E', 'G', 'I']

    # Iterate over the row and column ranges and merge cells
    for row in row_range:
        for col in letter_range:
            # Generate the range string for the current cell range
            range_string = f'{col}{row}:{chr(ord(col)+1)}{row}'

          # Merge the cells for the current range string
            ws.merge_cells(range_string=range_string)

    ws.merge_cells(range_string=f'A{max_temp+shift}:F{max_temp+shift}')
    ws.merge_cells(range_string=f'G{max_temp+shift}:L{max_temp+shift}')

    #Total Row Unmerge columns for all but last page
    ws.row_dimensions[shift-5].height = 30
    ws.unmerge_cells(f'A{shift-5}:B{shift-5}')
    ws.unmerge_cells(f'D{shift-5}:E{shift-5}')
    ws.unmerge_cells(f'G{shift-5}:H{shift-5}')
    ws.unmerge_cells(f'J{shift-5}:L{shift-5}')

  ################################################################################
  # convert the DataFrame to rows of data and append to the worksheet
  for r in dataframe_to_rows(df, index=True, header=False):
      ws.append(r)

  # get the range of cells that contain the appended data
  start_row = ws.max_row - len(df) + 1
  end_row = ws.max_row

  #Page iteration for DF #################################################################################
  for page in range(max_page):
    #Select Range for move
    shift = max_temp*page

    if page == max_page-1: cell_range = f'A{start_row+(page*(max_temp-10))}:K{end_row}' #if last page
    else: cell_range = f'A{start_row+(page*(max_temp-10))}:K{(start_row+(max_temp-11))+(page*(max_temp-10))}'

    # move the cell range to start at row 5
    ws.move_range(cell_range, rows=-(start_row+(page*(max_temp-10)))+(5+shift))

    #Format "bug" cell
    ws[f'G{(max_temp-2)+shift}'].border = Border(left=Side(style='medium'),
                                                right=Side(style='medium'),
                                                top=Side(style='thin'),
                                                bottom=Side(style='medium'))
    ws[f'H{(max_temp-2)+shift}'].border = Border(left=Side(style='medium'),
                                                right=Side(style='medium'),
                                                top=Side(style='thin'),
                                                bottom=Side(style='medium'))

    # define the range of rows to merge
    merge_start = 5+shift
    if page+1 != max_page:
      merge_end = (max_temp-5)+shift
      ws[f'G{(max_temp-4)+shift}'].border = Border(left=Side(style='medium'),
                                                  right=Side(style='medium'),
                                                  top=Side(style='medium'),
                                                  bottom=Side(style='medium'))
      ws[f'H{(max_temp-4)+shift}'].border = Border(left=Side(style='medium'),
                                                  right=Side(style='medium'),
                                                  top=Side(style='medium'),
                                                  bottom=Side(style='medium'))
      ws[f'I{(max_temp-4)+shift}'].border = Border(left=Side(style='medium'),
                                                  right=Side(style='thin'),
                                                  top=Side(style='medium'),
                                                  bottom=Side(style='medium'))
      ws[f'J{(max_temp-4)+shift}'].border = Border(left=Side(style='medium'),
                                                  right=Side(style='thin'),
                                                  top=Side(style='medium'),
                                                  bottom=Side(style='medium'))


    else:
      merge_end = ((max_temp-5)+shift)-1
      ws[f'J{(max_temp-5)+shift}'].border = Border(left=Side(style='medium'),
                                                  right=Side(style='thin'),
                                                  top=Side(style='medium'),
                                                  bottom=Side(style='medium'))
      ws[f'J{(max_temp-5)+shift}'].border = Border(left=Side(style='medium'),
                                                  right=Side(style='thin'),
                                                  top=Side(style='medium'),
                                                  bottom=Side(style='medium'))


  ############################################################################
  #DF FORMAT
  ###########################################################################

  # Create a border style
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))

    for row_num in range(merge_start, merge_end):
      ws.merge_cells(f'D{row_num}:E{row_num}')
      ws.merge_cells(f'K{row_num}:L{row_num}')
      ws.row_dimensions[row_num].height = 30

      for cell in ws[row_num]:
            if cell.column in [11,12]:
                cell.alignment = Alignment(vertical = 'center',
                                          horizontal = 'left',
                                          wrap_text = True)
                cell.border = Border(left=Side(style='thin'),
                                    right=Side(style='medium'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'))
                cell.font = Font(size=9)
            elif cell.column in [6,7,8,9,10]:
                cell.border = border_style
                cell.alignment = Alignment(vertical = 'center',
                                          horizontal = 'right')
            elif cell.column in [1,2,3,4,5]:
                cell.border = border_style
                cell.alignment = Alignment(vertical='center',
                                          horizontal = 'center')

  #Delete row after last footer (index header)
  ws.delete_rows(start_row-1)

  #Another iteration to add white space between df and footer
  for page in range(max_page):
    shift = max_temp*page
    const_row = (max_temp - 6) + shift
    if (page == max_page-1) and ws.cell(row=const_row, column=1).value is not None:
        ws.merge_cells(f'D{const_row}:E{const_row}')
        ws.merge_cells(f'K{const_row}:L{const_row}')
        ws.row_dimensions[const_row].height = 30
        for cell in ws[const_row]:
            if cell.column in [11,12]:
                cell.alignment = Alignment(vertical = 'center',
                                          horizontal = 'left',
                                          wrap_text = True)
                cell.border = Border(left=Side(style='thin'),
                                    right=Side(style='medium'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'))
                cell.font = Font(size=9)
            elif cell.column in [6,7,8,9,10]:
                cell.border = border_style
                cell.alignment = Alignment(vertical = 'center',
                                          horizontal = 'right')
            elif cell.column in [1,2,3,4,5]:
                cell.border = border_style
                cell.alignment = Alignment(vertical='center',
                                          horizontal = 'center')


    elif page == max_page-1: ws.row_dimensions[const_row].height = 9
    else:
      ws.row_dimensions[(max_temp-5)+shift].height = 9
      for cell in ws[(max_temp-5)+shift]:
        cell.value = None
        cell.border = Border()

  ################################################################################
  ################################################################################
  ################################################################################

  # Prepare the email content
  sender_email = os.environ["send_email"]                                         
  receiver_email = rec_email                                                     
  subject = "Here's the Excel file you requested"
  body = "Please find the attached Excel file."

  # Create a MIME multipart message to hold the email content and attachment
  msg = MIMEMultipart()
  msg['From'] = sender_email
  msg['To'] = receiver_email
  msg['Subject'] = subject

  # Attach the email body
  msg.attach(MIMEText(body, 'plain'))

  # Save the workbook to a temporary file
  temp_file_path = "/tmp/temp.xlsx"
  wb.save(temp_file_path)

  # Attach the temporary Excel file as an attachment
  with open(temp_file_path, "rb") as f:
      part = MIMEApplication(f.read(), Name="Reporte.xlsx")
      part['Content-Disposition'] = f'attachment; filename="Reporte.xlsx"'        
      msg.attach(part)


  # Send the email using a SMTP server (Gmail)
  smtp_server = "smtp.gmail.com"
  smtp_port = 587

  # Replace 'your_gmail_email' and 'your_gmail_password' with your actual Gmail credentials
  smtp_username = sender_email
  smtp_password = os.environ["smtp_pass"]

  with smtplib.SMTP(smtp_server, smtp_port) as server:
      server.starttls()
      # Authenticate with Gmail credentials
      server.login(smtp_username, smtp_password)
      server.sendmail(sender_email, receiver_email, msg.as_string())

  os.remove(temp_file_path)

  print("Email sent successfully.")

  return {
      'statusCode': 200,
      'body': 'Function executed successfully'
      }


